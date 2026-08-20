"""
Headless self-check for the packaged app.

    LogValidation.exe --selftest

Builds a small synthetic workbook in memory, runs all four programs against it,
and checks four files came out. Everything it needs is generated here, so it
works on a machine that has never seen the project data.

This exists because the failures PyInstaller causes are the ones that do not
show up until someone double-clicks the exe: a missing reference CSV, openpyxl's
package data left behind, tkinter not bundled. Those should surface in a build
step, not on site.
"""

from __future__ import annotations

import io
import sys
import tempfile
from pathlib import Path

APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))


def _project_des_rows() -> list[list]:
    """A miniature Project-Des.: the same blocks, side by side, one header row."""
    width = 27
    grid = [[None] * width for _ in range(9)]

    def put(row: int, col: int, *values):
        for offset, value in enumerate(values):
            grid[row][col + offset] = value

    put(0, 0, "Package")
    put(1, 0, "P08")

    put(0, 2, "P08")
    put(1, 2, "A1-008")
    put(2, 2, "C1-010")

    put(0, 4, "Tower_Floors", "Buidling_Floors", "Name")
    for index, (code, name) in enumerate(
        [("BS", "Basement"), ("GF", "Ground Floor"), ("1F", "1st Floor")], start=1
    ):
        put(index, 4, code, code, name)

    put(0, 8, "Code", "Apartment", "Package", "Building", "Floor", "Apt01", "Apt02")
    put(1, 8, "P08-A1-008-GF", "G01, G02", "P08", "A1-008", "GF", "G01", "G02")
    put(2, 8, "P08-A1-008-1F", "101, 102", "P08", "A1-008", "1F", "101", "102")
    put(3, 8, "P08-C1-010-BS", "L01", "P08", "C1-010", "BS", "L01")

    put(0, 16, "Code", "Meaning", "Counts as approved")
    put(1, 16, "B", "Approved", "Y")
    put(2, 16, "C", "Rejected", "N")

    put(0, 20, "Area", "Floors", "Variant 1",
        "Type", "Inside apartment", "Not a location", "Vertical")
    put(1, 20, "Corridor", None, "Cooridor", "COMMON", None, None, None)
    put(2, 20, "Electrical Room", None, "Electric Room", "SERVICE", None, None, None)
    put(3, 20, "Kitchen", None, "Kitchens", "COMMON", "Y", None, None)
    put(4, 20, "All Apartments", None, None, "SCOPE", None, "Y", None)
    return grid


WIR_HEADERS = [
    "Sr", "Submittal Reference. No.", "Description of Submission", "Dep.",
    "Package", "Building", "Floor", "Area", "Activity", "Status", "Rev", "Link",
]
CLEARANCE_HEADERS = [
    "Sr. no", "Clearance no. - REV", "Clearance no.", "Rev.",
    "Clearance Description", "Package", "Building", "Level", "AREA", "Activity",
    "Electrical", "Final", "Incoming",
]


def build_workbook() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()

    project = workbook.active
    project.title = "Project-Des."
    for line in _project_des_rows():
        project.append(line)

    wirs = workbook.create_sheet("WIRs")
    wirs.append([None])
    wirs.append([None])
    wirs.append(WIR_HEADERS)
    wirs.append([
        1, "UAE044103-INB-SP-P08-CO-PRW-ELE-WIR-000101", "Inspection request", "ELE",
        "P08", "A1-008", "1F", "Apartment 101, 102", "Plaster_Dry", "B", 0,
        r"\\192.168.0.1\share\wir.pdf",
    ])
    wirs.append([
        2, "UAE044103-INB-SP-P08-CO-PRW-ELE-WIR-000102", "Inspection request", "ELE",
        "P08", "A1-008", "GF", "Corridor", "Plaster_Dry", "C", 0,
        r"\\192.168.0.1\share\wir2.pdf",
    ])

    clearances = workbook.create_sheet("MEP Clearances")
    clearances.append([None])
    clearances.append([None])
    clearances.append(CLEARANCE_HEADERS)
    clearances.append([
        1, "CL-MEP-0001-R00", "CL-MEP-0001", 0,
        "MEP Clearance to proceed with plaster works", "P08", "A1-008", "1F",
        "Apartment 101, 102", "Plaster_Dry Area", "Pending", "Pending",
        r"\\192.168.0.1\share\cl.pdf",
    ])
    clearances.append([
        2, "CL-MEP-0002-R00", "CL-MEP-0002", 0,
        "MEP Clearance to proceed with plaster works", "P08", "A1-008", "GF",
        "Electric Room", "Plaster_Dry Area", "Rejected", "Rejected",
        r"\\192.168.0.1\share\cl2.pdf",
    ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _check_import_path(say) -> bool:
    """Exercise the Excel path the two Import buttons use.

    The validation half of the app never touches COM, so a broken Excel binding
    used to pass every check and then fail the moment someone pressed Import.
    That is how the exe shipped missing win32timezone: pywin32 imports it the
    first time COM hands back a date, which is on the very first Workbooks.Open,
    so nothing before that point notices.
    """
    try:
        import win32com.client as com  # noqa: F401
        import win32timezone  # noqa: F401
    except ImportError as exc:
        say(f"FAILED: Excel automation is not usable in this build - {exc}")
        return False

    import datetime
    import tempfile

    from openpyxl import Workbook

    try:
        excel = com.DispatchEx("Excel.Application")
    except Exception as exc:  # noqa: BLE001 - Excel genuinely absent is not our bug
        say(f"  note: Excel is not installed here, so Import was not exercised ({exc})")
        return True

    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        with tempfile.TemporaryDirectory() as workspace:
            probe = Path(workspace) / "com-probe.xlsx"
            book = Workbook()
            # A real date, because converting one is what pulls in win32timezone.
            book.active["A1"] = datetime.datetime(2026, 8, 18, 12, 30)
            book.save(probe)

            opened = excel.Workbooks.Open(str(probe))
            try:
                value = opened.Worksheets(1).Cells(1, 1).Value
                if value is None:
                    say("FAILED: Excel opened the probe but returned no value")
                    return False
                say(f"  Excel automation OK (read back {value})")
            finally:
                opened.Close(False)
    except Exception as exc:  # noqa: BLE001 - this is exactly what we are testing
        say(f"FAILED: Excel automation raised - {type(exc).__name__}: {exc}")
        return False
    finally:
        excel.Quit()

    return _check_a_real_import(say)


def _check_a_real_import(say) -> bool:
    """Run an actual import, because probing COM was not enough.

    Two Import failures reached the user in a row. Neither was caught, because
    the check stopped at "can Excel be driven" — and both were about what
    happens once real cell values start moving: a lazily imported pywin32
    module, then a datetime.time that COM refuses to marshal. So the source
    used here carries a date, a time and a blank, and the import is run for
    real.
    """
    import datetime
    import tempfile

    from openpyxl import Workbook, load_workbook

    from engine.importer import ImportError_, import_log

    with tempfile.TemporaryDirectory() as workspace:
        master = Path(workspace) / "Master.xlsx"
        book = Workbook()
        book.active.title = "WIRs"
        book.save(master)

        source = Path(workspace) / "Source.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "ELE"
        sheet.append([None])
        sheet.append([None])
        sheet.append(["Sr", "Submittal Reference. No.", "Desc", "Excl", "Dep.",
                      "Package", "Building", "Floor", "Area", "Activity", "Code",
                      "LOCATION", "TPD", "Link", "Rev", "Status", "days",
                      "Submitted", "Received", "WF"])
        sheet.append([1, "UAE-WIR-000695", "Inspection", None, "ELE", "P08",
                      "A1-008", "1F", "Apartment 101", "Plaster_Dry", "P08", "LOC",
                      None, r"\\host\share\f.pdf", 0, "B",
                      datetime.time(9, 30),                 # refuses to marshal
                      datetime.datetime(2026, 5, 13),       # needs win32timezone
                      datetime.date(2026, 5, 20), "WF-1"])
        book.save(source)

        try:
            result = import_log(master, source, "wir")
        except ImportError_ as exc:
            say(f"FAILED: import raised - {exc}")
            return False
        except Exception as exc:  # noqa: BLE001 - the whole point of this check
            say(f"FAILED: import raised - {type(exc).__name__}: {exc}")
            return False

        check = load_workbook(master, data_only=True, read_only=True)
        try:
            row = list(check["WIRs"].iter_rows(min_row=4, max_row=4, values_only=True))[0]
        finally:
            check.close()

        if str(row[1]) != "UAE-WIR-000695":
            say(f"FAILED: import wrote the wrong data - {row[1]!r}")
            return False
        say(f"  Import OK ({result.rows} row, dates and times survived)")
    return True


def main() -> int:
    # A windowed exe has no console, so the build step checks the exit code and
    # reads this file when it wants to know what went wrong.
    log = _open_log()

    def say(message: str) -> None:
        print(message)
        if log:
            log.write(message + "\n")

    try:
        return _run(say)
    finally:
        if log:
            log.close()


def _open_log():
    try:
        from desktop.app import app_dir

        return (app_dir() / "selftest.log").open("w", encoding="utf-8")
    except Exception:  # noqa: BLE001 - the log is a convenience, never required
        return None


def _run(say) -> int:
    from engine.pipeline import PRODUCTS, run

    say("Self-check: building a synthetic workbook...")
    data = build_workbook()

    from desktop.app import bundled_reference

    reference = bundled_reference()
    say(f"  reference data: {reference}")
    if not (reference / "constants.csv").exists():
        say(f"  FAILED: reference CSVs are not bundled at {reference}")
        return 1

    with tempfile.TemporaryDirectory() as workspace:
        source = Path(workspace) / "Selftest.xlsm"
        source.write_bytes(data)
        output = Path(workspace) / "out"

        report = run(
            source_path=source,
            output_dir=output,
            config_dir=reference,
            progress=lambda message, level: say(f"  {message}"),
        )

        expected = {spec.filename for spec in PRODUCTS.values()}
        produced = {p.name for p in output.iterdir()} if output.exists() else set()
        missing = expected - produced

    say("")
    if not _check_import_path(say):
        return 1

    if report.failed_early:
        say(f"FAILED: {report.failed_early}")
        return 1
    for result in report.results:
        if not result.ok:
            say(f"FAILED: {result.spec.title}: {result.error}")
    if missing:
        say(f"FAILED: no output for {', '.join(sorted(missing))}")
        return 1
    if not report.ok:
        return 1

    say(f"Self-check passed - {len(produced)} workbooks generated.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

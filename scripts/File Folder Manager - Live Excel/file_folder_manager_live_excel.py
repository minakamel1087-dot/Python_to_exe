"""Standalone, safe replacement for the workbook file/folder manager."""

from __future__ import annotations

import csv
import logging
import os
import shutil
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from PySide6.QtCore import Qt
from PySide6.QtGui import QAction, QColor
from PySide6.QtWidgets import (
    QApplication, QAbstractItemView, QCheckBox, QDialog, QFileDialog, QFormLayout,
    QComboBox, QGridLayout, QHBoxLayout, QLabel, QLineEdit, QMainWindow, QMessageBox,
    QPushButton, QTabWidget, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget,
)
from pypdf import PdfWriter
from send2trash import send2trash

try:
    import win32com.client
except ImportError:  # Allows source review on machines without Excel automation.
    win32com = None


# Keep runtime data alongside the executable in a portable Windows build.  In
# source mode it remains inside the project folder.
APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
LOG_DB = DATA_DIR / "activity_log.sqlite3"
INVALID_NAME_CHARS = set('\\/:*?<>|[]"')


@dataclass
class Item:
    name: str
    path: Path
    parent: Path
    modified: datetime
    proposed_name: str = ""
    destination: str = ""
    subfolder: str = ""
    status: str = "Ready"


class ActivityLog:
    def __init__(self) -> None:
        self.connection = sqlite3.connect(LOG_DB)
        self.connection.execute(
            """CREATE TABLE IF NOT EXISTS activity_log (
            timestamp TEXT, username TEXT, action TEXT, item_path TEXT, result TEXT, details TEXT)"""
        )
        self.connection.commit()

    def add(self, action: str, item_path: str, result: str, details: str = "") -> None:
        self.connection.execute(
            "INSERT INTO activity_log VALUES (?, ?, ?, ?, ?, ?)",
            (datetime.now().isoformat(timespec="seconds"), os.getenv("USERNAME", "unknown"),
             action, item_path, result, details),
        )
        self.connection.commit()

    def latest(self) -> list[tuple[str, ...]]:
        return self.connection.execute(
            "SELECT timestamp, username, action, item_path, result, details "
            "FROM activity_log ORDER BY rowid DESC LIMIT 500"
        ).fetchall()


def valid_name(name: str) -> bool:
    return bool(name.strip()) and not any(char in name for char in INVALID_NAME_CHARS)


def unique_destination(destination_dir: Path, name: str) -> Path:
    candidate = destination_dir / name
    if not candidate.exists():
        return candidate
    suffix = candidate.suffix
    stem = candidate.stem
    number = 1
    while True:
        candidate = destination_dir / f"{stem} ({number}){suffix}"
        if not candidate.exists():
            return candidate
        number += 1


class ExcelBridge:
    """Reads and writes the currently open Excel workbook; never runs a macro."""

    FILE_SHEET = "FileManager"
    FOLDER_SHEET = "FolderManager"

    def __init__(self) -> None:
        self.excel = None
        self.workbook = None

    def connect(self) -> str:
        if win32com is None:
            raise RuntimeError("Excel automation is unavailable in this build.")
        try:
            self.excel = win32com.client.GetActiveObject("Excel.Application")
            self.workbook = self.excel.ActiveWorkbook
        except Exception as exc:
            raise RuntimeError("Open the target workbook in Microsoft Excel first.") from exc
        if self.workbook is None:
            raise RuntimeError("Microsoft Excel has no active workbook.")
        return str(self.workbook.Name)

    def sheet(self, kind: str):
        if self.workbook is None:
            self.connect()
        sheet_name = self.FILE_SHEET if kind == "file" else self.FOLDER_SHEET
        try:
            return self.workbook.Worksheets(sheet_name)
        except Exception as exc:
            raise RuntimeError(f"The active workbook does not contain a '{sheet_name}' sheet.") from exc

    @staticmethod
    def _last_row(sheet) -> int:
        return max(2, sheet.Cells(sheet.Rows.Count, 1).End(-4162).Row)  # xlUp

    @staticmethod
    def _set_status(sheet, row: int, column: int, text: str, colour: int | None = None) -> None:
        cell = sheet.Cells(row, column)
        cell.Value = text
        if colour is not None:
            cell.Interior.Color = colour

    @staticmethod
    def _set_link(sheet, row: int, column: int, path: Path, label: str = "Open") -> None:
        cell = sheet.Cells(row, column)
        try:
            cell.Hyperlinks.Delete()
        except Exception:
            pass
        sheet.Hyperlinks.Add(cell, str(path), "", "Open item", label)

    def clear_list(self, kind: str) -> None:
        sheet = self.sheet(kind)
        last = self._last_row(sheet)
        if last >= 3:
            sheet.Range(sheet.Cells(3, 1), sheet.Cells(last, 12)).ClearContents()

    def write_scan(self, kind: str, items: Iterable[Item]) -> int:
        sheet = self.sheet(kind)
        self.clear_list(kind)
        rows = list(items)
        if not rows:
            return 0
        values = []
        for item in rows:
            date_text = item.modified.strftime("%d/%m/%Y")
            time_text = item.modified.strftime("%H:%M:%S")
            if kind == "file":
                # A Name, B Link, C Parent, D NewName, E Destination/Subfolder,
                # F Status, I Date, J Time, K FolderPath, L FilePath.
                values.append((item.name, "Open", str(item.parent), "", "", "Ready", "", "", date_text,
                               time_text, str(item.parent), str(item.path)))
            else:
                # A Name, B Link, C NewName, D NewParent, E Status, F/G date/time,
                # H Parent, I old path, J new path, K depth-sort key.
                values.append((item.name, "Open", "", "", "Ready", date_text, time_text, str(item.parent),
                               str(item.path), str(item.path), len(str(item.parent)), ""))
        start = 3
        sheet.Range(sheet.Cells(start, 1), sheet.Cells(start + len(values) - 1, 12)).Value = tuple(values)
        link_column = 2
        path_column = 12 if kind == "file" else 9
        for index, item in enumerate(rows, start):
            self._set_link(sheet, index, link_column, item.path)
        return len(rows)

    def selected_rows(self, kind: str, selected_only: bool) -> list[int]:
        sheet = self.sheet(kind)
        last = self._last_row(sheet)
        if selected_only:
            try:
                selection = self.excel.Selection
                if selection.Worksheet.Name != sheet.Name:
                    raise RuntimeError(f"Select rows on the '{sheet.Name}' sheet first.")
                rows = range(max(3, selection.Row), min(last, selection.Row + selection.Rows.Count - 1) + 1)
            except AttributeError as exc:
                raise RuntimeError("Select one or more data rows in Excel first.") from exc
        else:
            rows = range(3, last + 1)
        return [row for row in rows if not sheet.Rows(row).Hidden and str(sheet.Cells(row, 1).Value or "").strip()]


class ExcelLiveTab(QWidget):
    """Excel is the live grid; this tab only supplies commands and scan input."""

    def __init__(self, kind: str, activity: ActivityLog) -> None:
        super().__init__()
        self.kind = kind
        self.activity = activity
        self.bridge = ExcelBridge()
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            f"Live Microsoft Excel mode — uses the open '{'FileManager' if self.kind == 'file' else 'FolderManager'}' sheet. "
            "No VBA macros are run. Edit values in Excel, then run an action here."
        ))
        connection = QHBoxLayout()
        connect = QPushButton("Connect to active Excel workbook")
        connect.clicked.connect(self.connect_excel)
        activate = QPushButton("Open manager sheet in Excel")
        activate.clicked.connect(self.activate_sheet)
        self.connection_label = QLabel("Not connected")
        connection.addWidget(connect)
        connection.addWidget(activate)
        connection.addWidget(self.connection_label, 1)
        layout.addLayout(connection)

        scan_row = QHBoxLayout()
        self.source = QLineEdit()
        self.source.setPlaceholderText(f"Choose a folder to import {self.kind} information into Excel")
        browse = QPushButton("Choose folder…")
        browse.clicked.connect(self.choose_source)
        self.recursive = QCheckBox("Include subfolders")
        self.recursive.setChecked(True)
        scan = QPushButton(f"Import {self.kind}s into Excel")
        scan.clicked.connect(self.scan_to_excel)
        clear = QPushButton("Clear Excel list")
        clear.clicked.connect(self.clear_excel_list)
        scan_row.addWidget(self.source, 1)
        scan_row.addWidget(browse)
        scan_row.addWidget(self.recursive)
        scan_row.addWidget(scan)
        scan_row.addWidget(clear)
        layout.addLayout(scan_row)

        action_row = QHBoxLayout()
        self.scope = QComboBox()
        self.scope.addItems(["Selected Excel rows", "All visible Excel rows"])
        rename = QPushButton("Rename from Excel values")
        rename.clicked.connect(self.rename_from_excel)
        copy = QPushButton("Copy from Excel values")
        copy.clicked.connect(lambda: self.transfer_from_excel("copy"))
        move = QPushButton("Move from Excel values")
        move.clicked.connect(lambda: self.transfer_from_excel("move"))
        delete = QPushButton("Delete selected / visible rows")
        delete.clicked.connect(self.delete_from_excel)
        action_row.addWidget(QLabel("Process:"))
        action_row.addWidget(self.scope)
        action_row.addWidget(rename)
        action_row.addWidget(copy)
        action_row.addWidget(move)
        action_row.addWidget(delete)
        layout.addLayout(action_row)

        self.destination_mode = QComboBox()
        if self.kind == "file":
            self.destination_mode.addItems([
                "One chosen destination", "Destination from column E", "Chosen destination + subfolder from column E",
            ])
            details = "Files: rename from column D. Copy/move uses a chosen folder or column E. Status is written to column F."
        else:
            self.destination_mode.addItems(["One chosen destination", "New parent folder from column D"])
            details = "Folders: rename from column C; column D is an optional new parent path. Status is written to column E."
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("Copy/move destination:"))
        mode_row.addWidget(self.destination_mode)
        mode_row.addStretch()
        layout.addLayout(mode_row)
        layout.addWidget(QLabel(details))
        self.summary = QLabel("Connect to Microsoft Excel to begin.")
        layout.addWidget(self.summary)
        layout.addStretch()

    def connect_excel(self) -> bool:
        try:
            name = self.bridge.connect()
            self.bridge.sheet(self.kind)
            self.connection_label.setText(f"Connected: {name}")
            self.summary.setText("Connected. You can now import data or process edited Excel rows.")
            return True
        except RuntimeError as exc:
            QMessageBox.warning(self, "Excel connection", str(exc))
            return False

    def ensure_connection(self) -> bool:
        return self.connect_excel() if self.bridge.workbook is None else True

    def activate_sheet(self) -> None:
        if self.ensure_connection():
            self.bridge.sheet(self.kind).Activate()

    def choose_source(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Choose folder", self.source.text() or str(Path.home()))
        if folder:
            self.source.setText(folder)

    def scan_to_excel(self) -> None:
        if not self.ensure_connection():
            return
        root = Path(self.source.text().strip())
        if not root.is_dir():
            QMessageBox.warning(self, "Invalid folder", "Choose an existing folder to import.")
            return
        try:
            items = list(self._scan(root))
            count = self.bridge.write_scan(self.kind, items)
            self.activity.add(f"Import {self.kind}s to Excel", str(root), "Success", f"{count} item(s)")
            self.summary.setText(f"Imported {count} {self.kind}(s) into the open Excel workbook.")
            self.bridge.sheet(self.kind).Activate()
        except (OSError, RuntimeError) as exc:
            self.activity.add(f"Import {self.kind}s to Excel", str(root), "Failed", str(exc))
            QMessageBox.critical(self, "Import failed", str(exc))

    def _scan(self, root: Path) -> Iterable[Item]:
        for directory, folders, files in os.walk(root):
            names = files if self.kind == "file" else folders
            for name in names:
                path = Path(directory) / name
                try:
                    yield Item(name, path, path.parent, datetime.fromtimestamp(path.stat().st_mtime))
                except OSError as exc:
                    self.activity.add(f"Scan {self.kind}", str(path), "Skipped", str(exc))
            if not self.recursive.isChecked():
                break

    def clear_excel_list(self) -> None:
        if not self.ensure_connection():
            return
        if QMessageBox.question(self, "Clear Excel list", "Clear imported rows from the manager sheet?") != QMessageBox.StandardButton.Yes:
            return
        self.bridge.clear_list(self.kind)
        self.activity.add(f"Clear Excel {self.kind} list", self.bridge.sheet(self.kind).Name, "Success")
        self.summary.setText("Excel list cleared.")

    def _rows(self) -> list[int] | None:
        if not self.ensure_connection():
            return None
        try:
            rows = self.bridge.selected_rows(self.kind, self.scope.currentIndex() == 0)
            if not rows:
                raise RuntimeError("There are no matching Excel rows to process.")
            return rows
        except RuntimeError as exc:
            QMessageBox.information(self, "Select Excel rows", str(exc))
            return None

    def rename_from_excel(self) -> None:
        rows = self._rows()
        if not rows:
            return
        sheet = self.bridge.sheet(self.kind)
        succeeded = failed = 0
        # Rename deepest folders first so moving a parent cannot invalidate a child row.
        if self.kind == "folder":
            rows.sort(key=lambda row: len(str(sheet.Cells(row, 9).Value or "")), reverse=True)
        for row in rows:
            if self.kind == "file":
                source = Path(str(sheet.Cells(row, 12).Value or ""))
                new_name = str(sheet.Cells(row, 4).Value or "").strip()
                parent = source.parent
                status_col = 6
            else:
                source = Path(str(sheet.Cells(row, 9).Value or ""))
                new_name = str(sheet.Cells(row, 3).Value or "").strip()
                parent = Path(str(sheet.Cells(row, 4).Value or "").strip() or str(source.parent))
                status_col = 5
            try:
                if not valid_name(new_name):
                    raise ValueError("New name is empty or contains invalid characters")
                if not source.exists():
                    raise FileNotFoundError("Source path does not exist")
                if not parent.is_dir():
                    raise NotADirectoryError("Destination parent folder does not exist")
                target = parent / new_name
                if target.exists() and target != source:
                    raise FileExistsError("A file or folder with the new name already exists")
                source.rename(target)
                self._write_updated_path(sheet, row, target)
                self.bridge._set_status(sheet, row, status_col, "Renamed", 65280)
                self.activity.add(f"Rename {self.kind}", str(source), "Success", str(target))
                succeeded += 1
            except (OSError, ValueError) as exc:
                self.bridge._set_status(sheet, row, status_col, f"Failed: {exc}", 255)
                self.activity.add(f"Rename {self.kind}", str(source), "Failed", str(exc))
                failed += 1
        self.summary.setText(f"Rename complete — {succeeded} succeeded, {failed} failed. Excel has been updated.")

    def transfer_from_excel(self, operation: str) -> None:
        rows = self._rows()
        if not rows:
            return
        mode = self.destination_mode.currentIndex()
        base = ""
        if mode == 0 or (self.kind == "file" and mode == 2):
            base = QFileDialog.getExistingDirectory(self, f"Choose {operation} destination")
            if not base:
                return
        sheet = self.bridge.sheet(self.kind)
        status_col = 6 if self.kind == "file" else 5
        succeeded = failed = 0
        for row in rows:
            source_col = 12 if self.kind == "file" else 9
            source = Path(str(sheet.Cells(row, source_col).Value or ""))
            try:
                if self.kind == "file":
                    entry = str(sheet.Cells(row, 5).Value or "").strip()
                    if mode == 1:
                        target_dir = Path(entry)
                    elif mode == 2:
                        if not valid_name(entry):
                            raise ValueError("Column E must contain a valid subfolder name")
                        target_dir = Path(base) / entry
                    else:
                        target_dir = Path(base)
                else:
                    target_dir = Path(str(sheet.Cells(row, 4).Value or "").strip()) if mode == 1 else Path(base)
                if not source.exists():
                    raise FileNotFoundError("Source path does not exist")
                target_dir.mkdir(parents=True, exist_ok=True)
                target = unique_destination(target_dir, source.name)
                if operation == "copy":
                    shutil.copy2(source, target) if self.kind == "file" else shutil.copytree(source, target)
                    self.bridge._set_status(sheet, row, status_col, f"Copied: {target}", 65280)
                else:
                    shutil.move(str(source), str(target))
                    self._write_updated_path(sheet, row, target)
                    self.bridge._set_status(sheet, row, status_col, "Moved", 65280)
                self.activity.add(f"{operation.title()} {self.kind}", str(source), "Success", str(target))
                succeeded += 1
            except (OSError, ValueError) as exc:
                self.bridge._set_status(sheet, row, status_col, f"Failed: {exc}", 255)
                self.activity.add(f"{operation.title()} {self.kind}", str(source), "Failed", str(exc))
                failed += 1
        self.summary.setText(f"{operation.title()} complete — {succeeded} succeeded, {failed} failed. Excel has been updated.")

    def delete_from_excel(self) -> None:
        rows = self._rows()
        if not rows:
            return
        if QMessageBox.question(
            self, "Send to recycle bin", f"Send {len(rows)} {self.kind}(s) to the Windows recycle bin?"
        ) != QMessageBox.StandardButton.Yes:
            return
        sheet = self.bridge.sheet(self.kind)
        source_col, status_col = (12, 6) if self.kind == "file" else (9, 5)
        succeeded = failed = 0
        for row in rows:
            source = Path(str(sheet.Cells(row, source_col).Value or ""))
            try:
                if not source.exists():
                    raise FileNotFoundError("Source path does not exist")
                send2trash(str(source))
                self.bridge._set_status(sheet, row, status_col, "Sent to recycle bin", 65535)
                self.activity.add(f"Delete {self.kind}", str(source), "Recycled")
                succeeded += 1
            except OSError as exc:
                self.bridge._set_status(sheet, row, status_col, f"Failed: {exc}", 255)
                self.activity.add(f"Delete {self.kind}", str(source), "Failed", str(exc))
                failed += 1
        self.summary.setText(f"Delete complete — {succeeded} recycled, {failed} failed. Excel has been updated.")

    def _write_updated_path(self, sheet, row: int, target: Path) -> None:
        if self.kind == "file":
            sheet.Cells(row, 1).Value = target.name
            sheet.Cells(row, 3).Value = str(target.parent)
            sheet.Cells(row, 11).Value = str(target.parent)
            sheet.Cells(row, 12).Value = str(target)
            self.bridge._set_link(sheet, row, 2, target)
        else:
            sheet.Cells(row, 1).Value = target.name
            sheet.Cells(row, 8).Value = str(target.parent)
            sheet.Cells(row, 9).Value = str(target)
            sheet.Cells(row, 10).Value = str(target)
            self.bridge._set_link(sheet, row, 2, target)


class ItemTab(QWidget):
    """File/folder scan and batch-operation screen."""

    columns = ["Name", "Open path", "Parent", "New name", "Destination", "Subfolder", "Modified", "Status"]

    def __init__(self, kind: str, activity: ActivityLog) -> None:
        super().__init__()
        self.kind = kind
        self.activity = activity
        self.items: list[Item] = []
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        source_row = QHBoxLayout()
        self.source = QLineEdit()
        self.source.setPlaceholderText(f"Choose a root folder to scan for {self.kind}s")
        choose = QPushButton("Choose folder…")
        choose.clicked.connect(self.choose_source)
        self.recursive = QCheckBox("Include subfolders")
        self.recursive.setChecked(True)
        scan = QPushButton(f"Scan {self.kind}s")
        scan.clicked.connect(self.scan)
        source_row.addWidget(self.source, 1)
        source_row.addWidget(choose)
        source_row.addWidget(self.recursive)
        source_row.addWidget(scan)
        layout.addLayout(source_row)

        actions = QHBoxLayout()
        self.filter_box = QLineEdit()
        self.filter_box.setPlaceholderText("Filter current list")
        self.filter_box.textChanged.connect(self.apply_filter)
        for label, callback in [
            ("Rename selected", self.rename_selected), ("Copy selected", lambda: self.transfer_selected("copy")),
            ("Move selected", lambda: self.transfer_selected("move")), ("Delete selected", self.delete_selected),
            ("Export Excel…", self.export_excel), ("Clear", self.clear),
        ]:
            button = QPushButton(label)
            button.clicked.connect(callback)
            actions.addWidget(button)
        actions.insertWidget(0, self.filter_box, 1)
        layout.addLayout(actions)

        self.table = QTableWidget(0, len(self.columns))
        self.table.setHorizontalHeaderLabels(self.columns)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked | QAbstractItemView.EditTrigger.EditKeyPressed)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.cellDoubleClicked.connect(self.open_path)
        layout.addWidget(self.table)
        self.summary = QLabel("No items loaded")
        layout.addWidget(self.summary)

    def choose_source(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Choose folder", self.source.text() or str(Path.home()))
        if folder:
            self.source.setText(folder)

    def scan(self) -> None:
        root = Path(self.source.text().strip())
        if not root.is_dir():
            QMessageBox.warning(self, "Invalid folder", "Choose an existing folder to scan.")
            return
        self.items.clear()
        try:
            if self.kind == "file":
                iterator = self._scan_files(root)
            else:
                iterator = self._scan_folders(root)
            self.items = list(iterator)
            self.activity.add(f"Scan {self.kind}s", str(root), "Success", f"{len(self.items)} item(s)")
            self.populate()
        except OSError as exc:
            self.activity.add(f"Scan {self.kind}s", str(root), "Failed", str(exc))
            QMessageBox.critical(self, "Scan failed", str(exc))

    def _scan_files(self, root: Path) -> Iterable[Item]:
        walker = os.walk(root)
        for directory, _, files in walker:
            for name in files:
                path = Path(directory) / name
                try:
                    yield Item(name, path, path.parent, datetime.fromtimestamp(path.stat().st_mtime))
                except OSError as exc:
                    self.activity.add("Scan file", str(path), "Skipped", str(exc))
            if not self.recursive.isChecked():
                break

    def _scan_folders(self, root: Path) -> Iterable[Item]:
        walker = os.walk(root)
        for directory, folders, _ in walker:
            for name in folders:
                path = Path(directory) / name
                try:
                    yield Item(name, path, path.parent, datetime.fromtimestamp(path.stat().st_mtime))
                except OSError as exc:
                    self.activity.add("Scan folder", str(path), "Skipped", str(exc))
            if not self.recursive.isChecked():
                break

    def populate(self) -> None:
        self.table.setRowCount(len(self.items))
        for row, item in enumerate(self.items):
            values = [item.name, str(item.path), str(item.parent), item.proposed_name, item.destination,
                      item.subfolder, item.modified.strftime("%d/%m/%Y %H:%M:%S"), item.status]
            for col, value in enumerate(values):
                cell = QTableWidgetItem(value)
                if col in (0, 1, 2, 6, 7):
                    cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, col, cell)
        self.apply_filter()

    def sync_edits(self) -> None:
        for row, item in enumerate(self.items):
            item.proposed_name = self.table.item(row, 3).text().strip()
            item.destination = self.table.item(row, 4).text().strip()
            item.subfolder = self.table.item(row, 5).text().strip()

    def selected(self) -> list[tuple[int, Item]]:
        rows = sorted({index.row() for index in self.table.selectedIndexes()})
        return [(row, self.items[row]) for row in rows]

    def ensure_selection(self) -> list[tuple[int, Item]] | None:
        selected = self.selected()
        if not selected:
            QMessageBox.information(self, "Select items", "Select one or more rows first.")
            return None
        self.sync_edits()
        return selected

    def rename_selected(self) -> None:
        selected = self.ensure_selection()
        if not selected:
            return
        for _, item in selected:
            new_name = item.proposed_name
            if not valid_name(new_name):
                item.status = "Invalid new name"
                continue
            destination = item.path.with_name(new_name)
            if destination.exists() and destination != item.path:
                item.status = "Name already exists"
                continue
            try:
                item.path.rename(destination)
                self.activity.add(f"Rename {self.kind}", str(item.path), "Success", str(destination))
                item.path, item.name, item.parent, item.status = destination, destination.name, destination.parent, "Renamed"
            except OSError as exc:
                self.activity.add(f"Rename {self.kind}", str(item.path), "Failed", str(exc))
                item.status = f"Failed: {exc}"
        self.populate()

    def transfer_selected(self, operation: str) -> None:
        selected = self.ensure_selection()
        if not selected:
            return
        default_destination = QFileDialog.getExistingDirectory(self, f"Choose {operation} destination")
        if not default_destination:
            return
        for _, item in selected:
            target_dir = Path(item.destination) if item.destination else Path(default_destination)
            if item.subfolder:
                if not valid_name(item.subfolder):
                    item.status = "Invalid subfolder"
                    continue
                target_dir = target_dir / item.subfolder
            try:
                target_dir.mkdir(parents=True, exist_ok=True)
                target = unique_destination(target_dir, item.name)
                if operation == "copy":
                    if self.kind == "file":
                        shutil.copy2(item.path, target)
                    else:
                        shutil.copytree(item.path, target)
                    item.status = f"Copied → {target}"
                else:
                    shutil.move(str(item.path), str(target))
                    item.path, item.name, item.parent = target, target.name, target.parent
                    item.status = f"Moved → {target}"
                self.activity.add(f"{operation.title()} {self.kind}", str(item.path), "Success", str(target))
            except OSError as exc:
                self.activity.add(f"{operation.title()} {self.kind}", str(item.path), "Failed", str(exc))
                item.status = f"Failed: {exc}"
        self.populate()

    def delete_selected(self) -> None:
        selected = self.ensure_selection()
        if not selected:
            return
        reply = QMessageBox.question(
            self, "Send to recycle bin", f"Send {len(selected)} selected {self.kind}(s) to the Windows recycle bin?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        for _, item in selected:
            try:
                send2trash(str(item.path))
                self.activity.add(f"Delete {self.kind}", str(item.path), "Recycled")
                item.status = "Sent to recycle bin"
            except OSError as exc:
                self.activity.add(f"Delete {self.kind}", str(item.path), "Failed", str(exc))
                item.status = f"Failed: {exc}"
        self.populate()

    def export_excel(self) -> None:
        if not self.items:
            QMessageBox.information(self, "Nothing to export", "Scan items first.")
            return
        filename, _ = QFileDialog.getSaveFileName(self, "Export results", f"{self.kind}_manager.xlsx", "Excel (*.xlsx)")
        if not filename:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"{self.kind.title()} Manager"
            sheet.append(self.columns)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0070C0")
            for item in self.items:
                sheet.append([item.name, str(item.path), str(item.parent), item.proposed_name, item.destination,
                              item.subfolder, item.modified, item.status])
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column in sheet.columns:
                sheet.column_dimensions[column[0].column_letter].width = min(70, max(14, max(len(str(c.value or "")) for c in column) + 2))
            workbook.save(filename)
            self.activity.add(f"Export {self.kind}s", filename, "Success", f"{len(self.items)} item(s)")
        except Exception as exc:  # Export must surface all writer errors.
            self.activity.add(f"Export {self.kind}s", filename, "Failed", str(exc))
            QMessageBox.critical(self, "Export failed", str(exc))

    def clear(self) -> None:
        self.items.clear()
        self.table.setRowCount(0)
        self.summary.setText("No items loaded")

    def apply_filter(self) -> None:
        phrase = self.filter_box.text().lower().strip()
        visible = 0
        for row, item in enumerate(self.items):
            match = not phrase or phrase in " ".join(map(str, (item.name, item.path, item.parent, item.status))).lower()
            self.table.setRowHidden(row, not match)
            visible += int(match)
        self.summary.setText(f"{visible} visible of {len(self.items)} {self.kind}(s)")

    def open_path(self, row: int, column: int) -> None:
        if column in (0, 1, 2) and row < len(self.items):
            path = self.items[row].path if column != 2 else self.items[row].parent
            if path.exists():
                os.startfile(path)  # Windows desktop application.


class PdfMergeTab(QWidget):
    def __init__(self, activity: ActivityLog) -> None:
        super().__init__()
        self.activity = activity
        self.files: list[str] = []
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Add PDFs in the required output order, then choose where to save the merged file."))
        self.table = QTableWidget(0, 1)
        self.table.setHorizontalHeaderLabels(["PDF files (output order)"])
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)
        actions = QHBoxLayout()
        for label, callback in [("Add PDFs…", self.add_files), ("Remove selected", self.remove_selected),
                                ("Move up", lambda: self.reorder(-1)), ("Move down", lambda: self.reorder(1)),
                                ("Merge PDFs…", self.merge)]:
            button = QPushButton(label)
            button.clicked.connect(callback)
            actions.addWidget(button)
        layout.addLayout(actions)

    def add_files(self) -> None:
        files, _ = QFileDialog.getOpenFileNames(self, "Choose PDFs", filter="PDF files (*.pdf)")
        self.files.extend(files)
        self.populate()

    def populate(self) -> None:
        self.table.setRowCount(len(self.files))
        for row, filename in enumerate(self.files):
            self.table.setItem(row, 0, QTableWidgetItem(filename))

    def remove_selected(self) -> None:
        for row in sorted({i.row() for i in self.table.selectedIndexes()}, reverse=True):
            del self.files[row]
        self.populate()

    def reorder(self, direction: int) -> None:
        row = self.table.currentRow()
        new_row = row + direction
        if 0 <= row < len(self.files) and 0 <= new_row < len(self.files):
            self.files[row], self.files[new_row] = self.files[new_row], self.files[row]
            self.populate()
            self.table.selectRow(new_row)

    def merge(self) -> None:
        if not self.files:
            QMessageBox.information(self, "No PDFs", "Add one or more PDF files first.")
            return
        output, _ = QFileDialog.getSaveFileName(self, "Save merged PDF", "Merged.pdf", "PDF files (*.pdf)")
        if not output:
            return
        writer = PdfWriter()
        try:
            for file in self.files:
                writer.append(file)
            with open(output, "wb") as stream:
                writer.write(stream)
            self.activity.add("Merge PDFs", output, "Success", f"{len(self.files)} file(s)")
            QMessageBox.information(self, "PDFs merged", f"Saved {len(self.files)} PDF(s) to:\n{output}")
        except Exception as exc:
            self.activity.add("Merge PDFs", output, "Failed", str(exc))
            QMessageBox.critical(self, "Merge failed", str(exc))
        finally:
            writer.close()


class LogTab(QWidget):
    def __init__(self, activity: ActivityLog) -> None:
        super().__init__()
        self.activity = activity
        layout = QVBoxLayout(self)
        refresh = QPushButton("Refresh log")
        refresh.clicked.connect(self.refresh)
        layout.addWidget(refresh)
        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["Timestamp", "User", "Action", "Path", "Result", "Details"])
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)
        self.refresh()

    def refresh(self) -> None:
        rows = self.activity.latest()
        self.table.setRowCount(len(rows))
        for row, values in enumerate(rows):
            for col, value in enumerate(values):
                self.table.setItem(row, col, QTableWidgetItem(str(value)))


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("File and Folder Manager")
        self.resize(1400, 800)
        self.activity = ActivityLog()
        tabs = QTabWidget()
        tabs.addTab(ExcelLiveTab("file", self.activity), "Files — Live Excel")
        tabs.addTab(ExcelLiveTab("folder", self.activity), "Folders — Live Excel")
        tabs.addTab(PdfMergeTab(self.activity), "PDF Merge")
        tabs.addTab(LogTab(self.activity), "Activity Log")
        self.setCentralWidget(tabs)
        about = QAction("About", self)
        about.triggered.connect(lambda: QMessageBox.information(
            self, "About", "File and Folder Manager\nLive Excel companion with no licence gate or expiry."))
        self.menuBar().addMenu("Help").addAction(about)


def main() -> None:
    logging.basicConfig(level=logging.INFO)
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()

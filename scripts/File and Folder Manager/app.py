"""Standalone, safe replacement for the workbook file/folder manager."""

from __future__ import annotations

import csv
import logging
import os
import shutil
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from PySide6.QtCore import Qt
from PySide6.QtGui import QAction, QColor
from PySide6.QtWidgets import (
    QApplication, QAbstractItemView, QCheckBox, QDialog, QFileDialog, QFormLayout,
    QGridLayout, QHBoxLayout, QLabel, QLineEdit, QMainWindow, QMessageBox,
    QPushButton, QTabWidget, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget,
)
from pypdf import PdfWriter
from send2trash import send2trash


APP_DIR = Path(__file__).resolve().parent
LOG_DB = APP_DIR / "activity_log.sqlite3"
INVALID_NAME_CHARS = set('\\/:*?<>|[]"')


@dataclass
class Item:
    name: str
    path: Path
    parent: Path
    modified: datetime
    proposed_name: str = ""
    destination: str = ""
    subfolder: str = ""
    status: str = "Ready"


class ActivityLog:
    def __init__(self) -> None:
        self.connection = sqlite3.connect(LOG_DB)
        self.connection.execute(
            """CREATE TABLE IF NOT EXISTS activity_log (
            timestamp TEXT, username TEXT, action TEXT, item_path TEXT, result TEXT, details TEXT)"""
        )
        self.connection.commit()

    def add(self, action: str, item_path: str, result: str, details: str = "") -> None:
        self.connection.execute(
            "INSERT INTO activity_log VALUES (?, ?, ?, ?, ?, ?)",
            (datetime.now().isoformat(timespec="seconds"), os.getenv("USERNAME", "unknown"),
             action, item_path, result, details),
        )
        self.connection.commit()

    def latest(self) -> list[tuple[str, ...]]:
        return self.connection.execute(
            "SELECT timestamp, username, action, item_path, result, details "
            "FROM activity_log ORDER BY rowid DESC LIMIT 500"
        ).fetchall()


def valid_name(name: str) -> bool:
    return bool(name.strip()) and not any(char in name for char in INVALID_NAME_CHARS)


def unique_destination(destination_dir: Path, name: str) -> Path:
    candidate = destination_dir / name
    if not candidate.exists():
        return candidate
    suffix = candidate.suffix
    stem = candidate.stem
    number = 1
    while True:
        candidate = destination_dir / f"{stem} ({number}){suffix}"
        if not candidate.exists():
            return candidate
        number += 1


class ItemTab(QWidget):
    """File/folder scan and batch-operation screen."""

    columns = ["Name", "Open path", "Parent", "New name", "Destination", "Subfolder", "Modified", "Status"]

    def __init__(self, kind: str, activity: ActivityLog) -> None:
        super().__init__()
        self.kind = kind
        self.activity = activity
        self.items: list[Item] = []
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        source_row = QHBoxLayout()
        self.source = QLineEdit()
        self.source.setPlaceholderText(f"Choose a root folder to scan for {self.kind}s")
        choose = QPushButton("Choose folder…")
        choose.clicked.connect(self.choose_source)
        self.recursive = QCheckBox("Include subfolders")
        self.recursive.setChecked(True)
        scan = QPushButton(f"Scan {self.kind}s")
        scan.clicked.connect(self.scan)
        source_row.addWidget(self.source, 1)
        source_row.addWidget(choose)
        source_row.addWidget(self.recursive)
        source_row.addWidget(scan)
        layout.addLayout(source_row)

        actions = QHBoxLayout()
        self.filter_box = QLineEdit()
        self.filter_box.setPlaceholderText("Filter current list")
        self.filter_box.textChanged.connect(self.apply_filter)
        for label, callback in [
            ("Rename selected", self.rename_selected), ("Copy selected", lambda: self.transfer_selected("copy")),
            ("Move selected", lambda: self.transfer_selected("move")), ("Delete selected", self.delete_selected),
            ("Export Excel…", self.export_excel), ("Clear", self.clear),
        ]:
            button = QPushButton(label)
            button.clicked.connect(callback)
            actions.addWidget(button)
        actions.insertWidget(0, self.filter_box, 1)
        layout.addLayout(actions)

        self.table = QTableWidget(0, len(self.columns))
        self.table.setHorizontalHeaderLabels(self.columns)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked | QAbstractItemView.EditTrigger.EditKeyPressed)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.cellDoubleClicked.connect(self.open_path)
        layout.addWidget(self.table)
        self.summary = QLabel("No items loaded")
        layout.addWidget(self.summary)

    def choose_source(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Choose folder", self.source.text() or str(Path.home()))
        if folder:
            self.source.setText(folder)

    def scan(self) -> None:
        root = Path(self.source.text().strip())
        if not root.is_dir():
            QMessageBox.warning(self, "Invalid folder", "Choose an existing folder to scan.")
            return
        self.items.clear()
        try:
            if self.kind == "file":
                iterator = self._scan_files(root)
            else:
                iterator = self._scan_folders(root)
            self.items = list(iterator)
            self.activity.add(f"Scan {self.kind}s", str(root), "Success", f"{len(self.items)} item(s)")
            self.populate()
        except OSError as exc:
            self.activity.add(f"Scan {self.kind}s", str(root), "Failed", str(exc))
            QMessageBox.critical(self, "Scan failed", str(exc))

    def _scan_files(self, root: Path) -> Iterable[Item]:
        walker = os.walk(root)
        for directory, _, files in walker:
            for name in files:
                path = Path(directory) / name
                try:
                    yield Item(name, path, path.parent, datetime.fromtimestamp(path.stat().st_mtime))
                except OSError as exc:
                    self.activity.add("Scan file", str(path), "Skipped", str(exc))
            if not self.recursive.isChecked():
                break

    def _scan_folders(self, root: Path) -> Iterable[Item]:
        walker = os.walk(root)
        for directory, folders, _ in walker:
            for name in folders:
                path = Path(directory) / name
                try:
                    yield Item(name, path, path.parent, datetime.fromtimestamp(path.stat().st_mtime))
                except OSError as exc:
                    self.activity.add("Scan folder", str(path), "Skipped", str(exc))
            if not self.recursive.isChecked():
                break

    def populate(self) -> None:
        self.table.setRowCount(len(self.items))
        for row, item in enumerate(self.items):
            values = [item.name, str(item.path), str(item.parent), item.proposed_name, item.destination,
                      item.subfolder, item.modified.strftime("%d/%m/%Y %H:%M:%S"), item.status]
            for col, value in enumerate(values):
                cell = QTableWidgetItem(value)
                if col in (0, 1, 2, 6, 7):
                    cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, col, cell)
        self.apply_filter()

    def sync_edits(self) -> None:
        for row, item in enumerate(self.items):
            item.proposed_name = self.table.item(row, 3).text().strip()
            item.destination = self.table.item(row, 4).text().strip()
            item.subfolder = self.table.item(row, 5).text().strip()

    def selected(self) -> list[tuple[int, Item]]:
        rows = sorted({index.row() for index in self.table.selectedIndexes()})
        return [(row, self.items[row]) for row in rows]

    def ensure_selection(self) -> list[tuple[int, Item]] | None:
        selected = self.selected()
        if not selected:
            QMessageBox.information(self, "Select items", "Select one or more rows first.")
            return None
        self.sync_edits()
        return selected

    def rename_selected(self) -> None:
        selected = self.ensure_selection()
        if not selected:
            return
        for _, item in selected:
            new_name = item.proposed_name
            if not valid_name(new_name):
                item.status = "Invalid new name"
                continue
            destination = item.path.with_name(new_name)
            if destination.exists() and destination != item.path:
                item.status = "Name already exists"
                continue
            try:
                item.path.rename(destination)
                self.activity.add(f"Rename {self.kind}", str(item.path), "Success", str(destination))
                item.path, item.name, item.parent, item.status = destination, destination.name, destination.parent, "Renamed"
            except OSError as exc:
                self.activity.add(f"Rename {self.kind}", str(item.path), "Failed", str(exc))
                item.status = f"Failed: {exc}"
        self.populate()

    def transfer_selected(self, operation: str) -> None:
        selected = self.ensure_selection()
        if not selected:
            return
        default_destination = QFileDialog.getExistingDirectory(self, f"Choose {operation} destination")
        if not default_destination:
            return
        for _, item in selected:
            target_dir = Path(item.destination) if item.destination else Path(default_destination)
            if item.subfolder:
                if not valid_name(item.subfolder):
                    item.status = "Invalid subfolder"
                    continue
                target_dir = target_dir / item.subfolder
            try:
                target_dir.mkdir(parents=True, exist_ok=True)
                target = unique_destination(target_dir, item.name)
                if operation == "copy":
                    if self.kind == "file":
                        shutil.copy2(item.path, target)
                    else:
                        shutil.copytree(item.path, target)
                    item.status = f"Copied → {target}"
                else:
                    shutil.move(str(item.path), str(target))
                    item.path, item.name, item.parent = target, target.name, target.parent
                    item.status = f"Moved → {target}"
                self.activity.add(f"{operation.title()} {self.kind}", str(item.path), "Success", str(target))
            except OSError as exc:
                self.activity.add(f"{operation.title()} {self.kind}", str(item.path), "Failed", str(exc))
                item.status = f"Failed: {exc}"
        self.populate()

    def delete_selected(self) -> None:
        selected = self.ensure_selection()
        if not selected:
            return
        reply = QMessageBox.question(
            self, "Send to recycle bin", f"Send {len(selected)} selected {self.kind}(s) to the Windows recycle bin?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        for _, item in selected:
            try:
                send2trash(str(item.path))
                self.activity.add(f"Delete {self.kind}", str(item.path), "Recycled")
                item.status = "Sent to recycle bin"
            except OSError as exc:
                self.activity.add(f"Delete {self.kind}", str(item.path), "Failed", str(exc))
                item.status = f"Failed: {exc}"
        self.populate()

    def export_excel(self) -> None:
        if not self.items:
            QMessageBox.information(self, "Nothing to export", "Scan items first.")
            return
        filename, _ = QFileDialog.getSaveFileName(self, "Export results", f"{self.kind}_manager.xlsx", "Excel (*.xlsx)")
        if not filename:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"{self.kind.title()} Manager"
            sheet.append(self.columns)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0070C0")
            for item in self.items:
                sheet.append([item.name, str(item.path), str(item.parent), item.proposed_name, item.destination,
                              item.subfolder, item.modified, item.status])
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column in sheet.columns:
                sheet.column_dimensions[column[0].column_letter].width = min(70, max(14, max(len(str(c.value or "")) for c in column) + 2))
            workbook.save(filename)
            self.activity.add(f"Export {self.kind}s", filename, "Success", f"{len(self.items)} item(s)")
        except Exception as exc:  # Export must surface all writer errors.
            self.activity.add(f"Export {self.kind}s", filename, "Failed", str(exc))
            QMessageBox.critical(self, "Export failed", str(exc))

    def clear(self) -> None:
        self.items.clear()
        self.table.setRowCount(0)
        self.summary.setText("No items loaded")

    def apply_filter(self) -> None:
        phrase = self.filter_box.text().lower().strip()
        visible = 0
        for row, item in enumerate(self.items):
            match = not phrase or phrase in " ".join(map(str, (item.name, item.path, item.parent, item.status))).lower()
            self.table.setRowHidden(row, not match)
            visible += int(match)
        self.summary.setText(f"{visible} visible of {len(self.items)} {self.kind}(s)")

    def open_path(self, row: int, column: int) -> None:
        if column in (0, 1, 2) and row < len(self.items):
            path = self.items[row].path if column != 2 else self.items[row].parent
            if path.exists():
                os.startfile(path)  # Windows desktop application.


class PdfMergeTab(QWidget):
    def __init__(self, activity: ActivityLog) -> None:
        super().__init__()
        self.activity = activity
        self.files: list[str] = []
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Add PDFs in the required output order, then choose where to save the merged file."))
        self.table = QTableWidget(0, 1)
        self.table.setHorizontalHeaderLabels(["PDF files (output order)"])
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)
        actions = QHBoxLayout()
        for label, callback in [("Add PDFs…", self.add_files), ("Remove selected", self.remove_selected),
                                ("Move up", lambda: self.reorder(-1)), ("Move down", lambda: self.reorder(1)),
                                ("Merge PDFs…", self.merge)]:
            button = QPushButton(label)
            button.clicked.connect(callback)
            actions.addWidget(button)
        layout.addLayout(actions)

    def add_files(self) -> None:
        files, _ = QFileDialog.getOpenFileNames(self, "Choose PDFs", filter="PDF files (*.pdf)")
        self.files.extend(files)
        self.populate()

    def populate(self) -> None:
        self.table.setRowCount(len(self.files))
        for row, filename in enumerate(self.files):
            self.table.setItem(row, 0, QTableWidgetItem(filename))

    def remove_selected(self) -> None:
        for row in sorted({i.row() for i in self.table.selectedIndexes()}, reverse=True):
            del self.files[row]
        self.populate()

    def reorder(self, direction: int) -> None:
        row = self.table.currentRow()
        new_row = row + direction
        if 0 <= row < len(self.files) and 0 <= new_row < len(self.files):
            self.files[row], self.files[new_row] = self.files[new_row], self.files[row]
            self.populate()
            self.table.selectRow(new_row)

    def merge(self) -> None:
        if not self.files:
            QMessageBox.information(self, "No PDFs", "Add one or more PDF files first.")
            return
        output, _ = QFileDialog.getSaveFileName(self, "Save merged PDF", "Merged.pdf", "PDF files (*.pdf)")
        if not output:
            return
        writer = PdfWriter()
        try:
            for file in self.files:
                writer.append(file)
            with open(output, "wb") as stream:
                writer.write(stream)
            self.activity.add("Merge PDFs", output, "Success", f"{len(self.files)} file(s)")
            QMessageBox.information(self, "PDFs merged", f"Saved {len(self.files)} PDF(s) to:\n{output}")
        except Exception as exc:
            self.activity.add("Merge PDFs", output, "Failed", str(exc))
            QMessageBox.critical(self, "Merge failed", str(exc))
        finally:
            writer.close()


class LogTab(QWidget):
    def __init__(self, activity: ActivityLog) -> None:
        super().__init__()
        self.activity = activity
        layout = QVBoxLayout(self)
        refresh = QPushButton("Refresh log")
        refresh.clicked.connect(self.refresh)
        layout.addWidget(refresh)
        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["Timestamp", "User", "Action", "Path", "Result", "Details"])
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)
        self.refresh()

    def refresh(self) -> None:
        rows = self.activity.latest()
        self.table.setRowCount(len(rows))
        for row, values in enumerate(rows):
            for col, value in enumerate(values):
                self.table.setItem(row, col, QTableWidgetItem(str(value)))


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("File and Folder Manager")
        self.resize(1400, 800)
        self.activity = ActivityLog()
        tabs = QTabWidget()
        tabs.addTab(ItemTab("file", self.activity), "Files")
        tabs.addTab(ItemTab("folder", self.activity), "Folders")
        tabs.addTab(PdfMergeTab(self.activity), "PDF Merge")
        tabs.addTab(LogTab(self.activity), "Activity Log")
        self.setCentralWidget(tabs)
        about = QAction("About", self)
        about.triggered.connect(lambda: QMessageBox.information(
            self, "About", "File and Folder Manager\nStandalone Python replacement with no licence gate or expiry."))
        self.menuBar().addMenu("Help").addAction(about)


def main() -> None:
    logging.basicConfig(level=logging.INFO)
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
